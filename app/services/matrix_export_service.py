"""Editable Excel export for the engineer's URS traceability matrix."""

from __future__ import annotations

from io import BytesIO
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.domain.models import AnalysisRun, DocumentVersion, ReviewPackage
from app.domain.ports import TraceabilityMatrixRow


class MatrixExportService:
    """Build a self-contained workbook; no formulas are needed for this export."""

    MATRIX_HEADERS = [
        "URS ID",
        "URS Text",
        "Priority",
        "Impact / Rationale",
        "Critical",
        "Assessment Status",
        "Status Definition",
        "Processing Status",
        "Design-Specification Citations (document, version, section, page, excerpt)",
        "LLM Mapping Rationale",
        "Potential Gap",
        "Suggested Reviewer Action",
        "Internal Audit Points",
        "Technical Exception",
    ]

    def build(
        self,
        *,
        review: ReviewPackage,
        run: AnalysisRun,
        rows: Iterable[TraceabilityMatrixRow],
        documents: Iterable[DocumentVersion],
    ) -> bytes:
        workbook = Workbook()
        matrix = workbook.active
        matrix.title = "URS Traceability Matrix"
        metadata = workbook.create_sheet("Report Metadata")

        self._write_matrix(matrix, rows)
        self._write_metadata(metadata, review, run, documents)

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    def _write_matrix(self, sheet, rows: Iterable[TraceabilityMatrixRow]) -> None:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(self.MATRIX_HEADERS))}1"
        header_fill = PatternFill("solid", fgColor="1F4E78")
        for column, value in enumerate(self.MATRIX_HEADERS, start=1):
            cell = sheet.cell(row=1, column=column, value=value)
            cell.font = Font(name="Arial", bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row_index, row in enumerate(rows, start=2):
            values = [
                row.requirement_code,
                row.requirement_text,
                row.priority or "",
                row.rationale_impact or "",
                "Yes" if row.is_critical else "No",
                row.design_status or "Assessment incomplete / technical exception",
                row.status_definition or "No candidate conclusion is available for this URS; human review is required.",
                row.analysis_status,
                self._format_evidence(row),
                row.rationale or "",
                row.gap or "",
                row.suggested_reviewer_action or "",
                self._format_audit_points(row),
                row.technical_error or "",
            ]
            for column, value in enumerate(values, start=1):
                cell = sheet.cell(row=row_index, column=column, value=value)
                cell.font = Font(name="Arial")
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        widths = [18, 46, 14, 34, 10, 22, 42, 18, 60, 48, 34, 34, 56, 34]
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width
        for row_index in range(2, sheet.max_row + 1):
            sheet.row_dimensions[row_index].height = 72

    @staticmethod
    def _format_evidence(row: TraceabilityMatrixRow) -> str:
        if not row.evidence:
            return ""
        return "\n\n".join(
            f"{item.document_title} v{item.version} | {item.section or 'Unsectioned'} | p.{item.page or '—'}\n{item.excerpt}"
            for item in row.evidence
        )

    @staticmethod
    def _format_audit_points(row: TraceabilityMatrixRow) -> str:
        if not row.audit_points:
            return ""
        values = []
        for point in row.audit_points:
            citations = ", ".join(
                f"{item.document_title} v{item.version} p.{item.page or '—'}" for item in point.evidence
            )
            values.append(
                f"{point.point_id}: {point.review_point}\n"
                f"Status: {point.design_status}\n"
                f"Rationale: {point.rationale}\n"
                f"Citations: {citations or 'None'}"
            )
        return "\n\n".join(values)

    @staticmethod
    def _write_metadata(
        sheet,
        review: ReviewPackage,
        run: AnalysisRun,
        documents: Iterable[DocumentVersion],
    ) -> None:
        sheet.column_dimensions["A"].width = 28
        sheet.column_dimensions["B"].width = 100
        values = [
            ("Report Name", review.name),
            ("Review Package ID", review.id),
            ("System", review.system),
            ("URS baseline ID", review.requirement_baseline_id),
            ("Generated At", run.completed_at or run.created_at),
            ("Analysis Run ID", run.id),
            ("Run Status", run.status),
        ]
        for document in documents:
            values.extend(
                [
                    ("Evidence Document", f"{document.document.title} v{document.version}"),
                    ("Evidence Filename", document.file_name or ""),
                    ("Evidence File Hash", document.file_hash or ""),
                ]
            )
        for row_index, (key, value) in enumerate(values, start=1):
            key_cell = sheet.cell(row=row_index, column=1, value=key)
            value_cell = sheet.cell(row=row_index, column=2, value=str(value or ""))
            key_cell.font = Font(name="Arial", bold=True)
            value_cell.font = Font(name="Arial")
            value_cell.alignment = Alignment(wrap_text=True, vertical="top")
